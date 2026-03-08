from config import ARRIVAL_HEADER, LEAVING_HEADER

import openpyxl
from dateutil.parser import parse
from datetime import datetime

def load_workbook(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    headers = [cell for cell in next(ws.iter_rows(values_only=True))]
    return wb, ws, headers

def is_valid_date(item):
    if isinstance(item, datetime):
        return True
    elif isinstance(item, str):
        try:
            parse(item)
            return True
        except (ValueError, TypeError):
            return False
    return False

def extract_valid_rows(ws, headers):
    data = []

    idx_arrival_date = headers.index(ARRIVAL_HEADER)
    idx_leaving_date = headers.index(LEAVING_HEADER)

    for row in ws.iter_rows(values_only=True):
        if is_valid_date(row[idx_arrival_date]) and is_valid_date(row[idx_leaving_date]):
            data.append(row)
    return data
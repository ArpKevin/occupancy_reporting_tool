from openpyxl.utils import get_column_letter

def insert_nights(nights_per_date_range, ws, wb, file_name, headers):
    header_name = "Nights in month"

    # Try to find the column that already has this header
    col_idx = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header_name:
            col_idx = col
            break
    
    # If not found, append a new column at the end
    if col_idx is None:
        col_idx = ws.max_column + 1
        ws.cell(row=1, column=col_idx, value=header_name)

    for row_idx, nights in enumerate(nights_per_date_range, start=2):
        ws.cell(row=row_idx, column=col_idx, value=nights)

    wb.save(file_name)
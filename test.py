import openpyxl
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from file_utils.validator import get_file_name_cli

def load_workbook(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    return wb, ws

def is_valid_date(item):
    if isinstance(item, datetime):
        return True
    elif isinstance(item, str):
        try:
            datetime.strptime(item, '%m/%d/%Y')
            return True
        except ValueError:
            return False
    return False

def extract_valid_rows(ws):
    data = []
    for row in ws.iter_rows(values_only=True):
        if any(is_valid_date(item) for item in row):
            data.append(row)
    return data

def get_user_year():
    while True:
        try:
            year = int(input("Enter the year (YYYY): "))
            if 10000 > year > 999:
                return year
            else:
                print("Please enter a number between 1000 and 9999.")
        except ValueError:
            print("Please enter a valid number.")

def get_user_month():
    while True:
        try:
            month = int(input("Enter the month (MM): "))
            if 1 <= month <= 12:
                return month
            else:
                print("Please enter a number between 1 and 12.")
        except ValueError:
            print("Please enter a valid number.")

def calculate_nights_per_date_range(year, month, data):
    nights_per_date_range = []

    for row in data:
        arrival_date = row[1]
        leaving_date = row[2]

        if leaving_date >= datetime(year, month, 1) and arrival_date <= datetime(year, month, 1) + relativedelta(months=1, days=-1):
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year+1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(year, month+1, 1) - timedelta(days=1)

            nights = (min(end_date, leaving_date) - max(start_date, arrival_date)).days + 1

            if leaving_date.month == month and leaving_date.year == year:
                nights -= 1

            key = f'{arrival_date.date()} - {leaving_date.date()}'

            if 'Owner' in row or 'Čiščenje / hišnik' in row:
                nights = 0
                key += ' (Owner)' if 'Owner' in row else ' (Čiščenje / hišnik)'

            nights_per_date_range.append((key, nights))

        else:
            key = f'{arrival_date.date()} - {leaving_date.date()}'

            if 'Owner' in row or 'Čiščenje / hišnik' in row:
                key += ' (Owner)' if 'Owner' in row else ' (Čiščenje / hišnik)'
            
            nights_per_date_range.append((key, 0))
    return nights_per_date_range

def insert_nights(year, month, nights_per_date_range, ws, wb, file_name):
    input("\nPress \"Enter\" to insert the values into the xlsx...")

    print("Please wait. This may take a few seconds...")

    month_name = "January" if month == 1 else \
                "February" if month == 2 else \
                "March" if month == 3 else \
                "April" if month == 4 else \
                "May" if month == 5 else \
                "June" if month == 6 else \
                "July" if month == 7 else \
                "August" if month == 8 else \
                "September" if month == 9 else \
                "October" if month == 10 else \
                "November" if month == 11 else \
                "December"

    ws["K1"].value = f"Nights in month ({year} {month_name})"

    x = 2
    for date_range, nights in nights_per_date_range:
        ws[f"K{x}"].value = nights
        x+=1

    wb.save(file_name)

    input("\nData inserted successfully.")

def main():
    file_name = get_file_name_cli()
    wb, ws = load_workbook(file_name)
    data = extract_valid_rows(ws)
    year = get_user_year()
    month = get_user_month()
    nights_per_date_range = calculate_nights_per_date_range(year, month, data)
    insert_nights(year, month, nights_per_date_range, ws, wb, file_name)

if __name__ == "__main__":
    main()